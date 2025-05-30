from django.test import TestCase, Client
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO
from datetime import date
from apps.stock.models import Producto, Categoria, Subcategoria
from apps.empresa.models import Almacen, Sucursal, Empresa


class ExportarProductosExcelTest(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(
            nombre="Mi empresa", cuit="20123456789", direccion="Calle 123"
        )
        self.sucursal = Sucursal.objects.create(
            nombre="Sucursal Central", empresa=self.empresa, direccion="Av. Siempre Viva"
        )
        self.almacen = Almacen.objects.create(nombre="Depósito Central", sucursal=self.sucursal)

        self.categoria = Categoria.objects.create(nombre="Bebidas")
        self.subcategoria = Subcategoria.objects.create(nombre="Gaseosas", categoria=self.categoria)

        Producto.objects.create(
            codigo="PROD001",
            nombre="Coca Cola 1.5L",
            descripcion="Gaseosa cola",
            unidad="botella",
            precio=150.00,
            vencimiento=date.today(),
            stock=100,
            stock_minimo=10,
            categoria=self.categoria,
            subcategoria=self.subcategoria,
            almacen=self.almacen
        )

        self.client = Client()
        self.url = reverse('exportar_productos_excel')

    def test_exportacion_excel_respuesta_correcta(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('productos.xlsx', response['Content-Disposition'])

    def test_exportacion_excel_contenido_correcto(self):
        response = self.client.get(self.url)
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        encabezados_esperados = [
            "Código", "Nombre", "Precio", "Cantidad", "Unidad",
            "Categoría", "Subcategoría", "Almacén"
        ]
        encabezados_archivo = [cell.value for cell in ws[1]]
        self.assertEqual(encabezados_archivo, encabezados_esperados)

        fila = [cell.value for cell in ws[2]]
        self.assertEqual(fila[0], "prod001")
        self.assertEqual(fila[1], "coca cola 1.5l")
        self.assertEqual(fila[2], 150.00)
        self.assertEqual(fila[3], 100)
        self.assertEqual(fila[4], "botella")
        self.assertEqual(fila[5], "Bebidas")
        self.assertEqual(fila[6], "Gaseosas")
        self.assertEqual(fila[7], "Depósito Central")

    def test_exportacion_excel_sin_productos(self):
        Producto.objects.all().delete()

        response = self.client.get(self.url)
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        encabezados = [cell.value for cell in ws[1]]
        self.assertEqual(encabezados, [
            "Código", "Nombre", "Precio", "Cantidad", "Unidad",
            "Categoría", "Subcategoría", "Almacén"
        ])
        self.assertEqual(ws.max_row, 1)

    def test_exportar_producto_sin_subcategoria(self):
        Producto.objects.all().delete()

        Producto.objects.create(
            codigo="SINCAT001",
            nombre="Producto sin subcategoría",
            descripcion="Sin subcat",
            unidad="unidad",
            precio=100.00,
            vencimiento=date.today(),
            stock=20,
            stock_minimo=5,
            categoria=self.categoria,
            subcategoria=None,
            almacen=self.almacen
        )

        response = self.client.get(self.url)
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        fila = [cell.value for cell in ws[2]]
        self.assertEqual(fila[0], "sincat001")
        self.assertEqual(fila[6], "")  # Subcategoría vacía
