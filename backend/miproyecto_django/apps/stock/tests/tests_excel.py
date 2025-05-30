from django.test import TestCase
from rest_framework.test import APIClient
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from apps.usuarios.models import Usuario
from apps.empresa.models import Almacen

class CargaMasivaProductosTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = Usuario.objects.create_user(email='test@test.com', password='1234', rol='admin')
        self.client.force_authenticate(user=self.user)
        self.almacen = Almacen.objects.create(nombre='central', sucursal_id=1)

    def crear_excel_valido(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'codigo', 'nombre', 'descripcion', 'unidad',
            'precio', 'stock', 'vencimiento',
            'almacen', 'categoria', 'subcategoria'
        ])
        ws.append([
            'A001', 'agua', 'botella 1L', 'unidad',
            100.00, 10, '2025-12-31',
            'central', 'bebidas', 'sin gas'
        ])
        from io import BytesIO
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        return SimpleUploadedFile("productos.xlsx", archivo.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_carga_excel_valido_crea_producto(self):
        archivo = self.crear_excel_valido()
        response = self.client.post('/api/v1/stock/carga-masiva/', {'file': archivo}, format='multipart')
        self.assertEqual(response.status_code, 200)
        self.assertIn('productos_creados', response.data)
        self.assertEqual(response.data['productos_creados'], 1)

    def test_excel_faltante_columna(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['codigo', 'nombre'])  # columnas incompletas
        from io import BytesIO
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        file = SimpleUploadedFile("mal.xlsx", archivo.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response = self.client.post('/api/v1/stock/carga-masiva/', {'file': file}, format='multipart')
        self.assertEqual(response.status_code, 400)
        self.assertIn('Falta la columna obligatoria', response.data['error'])

    def test_excel_dato_invalido(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'codigo', 'nombre', 'descripcion', 'unidad',
            'precio', 'stock', 'vencimiento',
            'almacen', 'categoria', 'subcategoria'
        ])
        ws.append([
            'A002', 'agua', 'botella', 'unidad',
            'gratis', 10, '2025-12-31', 'central', 'bebidas', 'sin gas'
        ])
        from io import BytesIO
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        file = SimpleUploadedFile("error.xlsx", archivo.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response = self.client.post('/api/v1/stock/carga-masiva/', {'file': file}, format='multipart')
        self.assertEqual(response.status_code, 200)
        self.assertGreater(len(response.data['errores']), 0)

    def test_actualiza_stock_si_producto_ya_existe(self):
        # Crear primero
        archivo
