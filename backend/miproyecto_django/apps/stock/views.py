from django.shortcuts import render
from rest_framework import viewsets
from .serializers import CategoriaSerializer, SubcategoriaSerializer, ProductoSerializer, MovimientoStockSerializer
from rest_framework.permissions import IsAuthenticated
import pandas as pd
from unidecode import unidecode
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from .models import Producto, Categoria, Subcategoria, MovimientoStock
from apps.empresa.models import Almacen
from apps.proveedores.models import Proveedor, Compra
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ProductoFilter
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Producto


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = ProductoFilter
    permission_classes = [IsAuthenticated]


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated]

class SubcategoriaViewSet(viewsets.ModelViewSet):
    queryset = Subcategoria.objects.all()
    serializer_class = SubcategoriaSerializer
    permission_classes = [IsAuthenticated]

class MovimientoStockViewSet(viewsets.ModelViewSet):
    queryset = MovimientoStock.objects.all()
    serializer_class = MovimientoStockSerializer
    permission_classes = [IsAuthenticated]


def limpiar_texto(valor):
    if pd.isnull(valor):
        return ''
    return unidecode(str(valor).strip().lower())

class CargaMasivaProductosView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        # Validación del usuario
        if not request.user.is_authenticated:
            return Response({"error": "Usuario no autenticado."}, status=403)

        archivo = request.FILES.get('file')
        if not archivo:
            return Response({"error": "No se envió ningún archivo."}, status=400)

        try:
            df = pd.read_excel(archivo)
        except Exception as e:
            return Response({"error": f"Error al leer el archivo Excel: {str(e)}"}, status=400)

        # Validar columnas obligatorias
        columnas_obligatorias = [
            'codigo', 'nombre', 'descripcion', 'unidad',
            'precio', 'stock', 'vencimiento',
            'almacen', 'categoria', 'subcategoria'
        ]
        for col in columnas_obligatorias:
            if col not in df.columns:
                return Response({"error": f"Falta la columna obligatoria: '{col}'"}, status=400)

        errores = []
        productos_creados = 0
        productos_actualizados = 0

        for i, row in df.iterrows():
            if row.dropna().empty:
                continue  # saltar filas vacías

            try:
                # Normalización de campos
                codigo = limpiar_texto(row['codigo'])
                nombre = limpiar_texto(row['nombre'])
                descripcion = str(row['descripcion']).strip()
                unidad = limpiar_texto(row['unidad'])
                precio = float(row['precio'])
                stock = int(row['stock'])
                vencimiento = pd.to_datetime(row['vencimiento'], errors='coerce')
                almacen_nombre = limpiar_texto(row['almacen'])
                categoria_nombre = limpiar_texto(row['categoria'])
                subcategoria_nombre = limpiar_texto(row['subcategoria'])

                if not codigo or not nombre or precio <= 0 or stock < 0:
                    raise ValueError("Datos obligatorios inválidos (código, nombre, precio, stock).")

                # Buscar o crear almacén, categoría y subcategoría
                almacen = Almacen.objects.filter(nombre__iexact=almacen_nombre).first()
                if not almacen:
                    raise ValueError(f"Almacén '{almacen_nombre}' no existe.")

                categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
                subcategoria, _ = Subcategoria.objects.get_or_create(nombre=subcategoria_nombre, categoria=categoria)

                producto, creado = Producto.objects.get_or_create(
                    codigo=codigo,
                    almacen=almacen,
                    defaults={
                        'nombre': nombre,
                        'descripcion': descripcion,
                        'unidad': unidad,
                        'precio': precio,
                        'stock': 0,
                        'vencimiento': vencimiento,
                        'categoria': categoria,
                        'subcategoria': subcategoria,
                    }
                )

                if creado:
                    producto.stock = stock
                    productos_creados += 1
                else:
                    producto.stock += stock
                    productos_actualizados += 1

                producto.save()

                # Registrar movimiento de stock
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='entrada',
                    cantidad=stock,
                    motivo='Carga masiva',
                    #usuario=request.user
                )

            except Exception as e:
                errores.append(f"Fila {i+2}: {str(e)}")  # +2 por encabezado de Excel

        return Response({
            "mensaje": "Carga finalizada",
            "productos_creados": productos_creados,
            "productos_actualizados": productos_actualizados,
            "errores": errores
        }, status=200)


def exportar_productos_excel(request):
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    columnas = ["Código", "Nombre", "Precio", "Cantidad", "Unidad", "Categoría", "Subcategoría", "Almacén"]
    for col_index, column_title in enumerate(columnas, 1):
        ws[f"{get_column_letter(col_index)}1"] = column_title

    # Datos
    productos = Producto.objects.select_related("categoria", "subcategoria", "almacen").all()
    for row_index, producto in enumerate(productos, start=2):
        ws[f"A{row_index}"] = producto.codigo
        ws[f"B{row_index}"] = producto.nombre
        ws[f"C{row_index}"] = float(producto.precio)
        ws[f"D{row_index}"] = producto.cantidad
        ws[f"E{row_index}"] = producto.unidad
        ws[f"F{row_index}"] = producto.categoria.nombre if producto.categoria else ""
        ws[f"G{row_index}"] = producto.subcategoria.nombre if producto.subcategoria else ""
        ws[f"H{row_index}"] = producto.almacen.nombre if producto.almacen else ""

    # Respuesta HTTP con archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=productos.xlsx"
    wb.save(response)
    return response




class PrevisualizarAjustePreciosView(APIView):
    def post(self, request):
        tipo = request.data.get("tipo")
        grupo_id = request.data.get("id")
        porcentaje = request.data.get("porcentaje")

        if tipo not in ["cat", "sub", "prov"]:
            return Response({"error": "Tipo debe ser 'cat', 'sub' o 'prov'."}, status=400)

        try:
            porcentaje = float(porcentaje)
        except (TypeError, ValueError):
            return Response({"error": "El porcentaje debe ser un número válido."}, status=400)

        # Buscar productos según el tipo
        if tipo == "cat":
            categoria = Categoria.objects.filter(id=grupo_id).first()
            if not categoria:
                return Response({"error": "Categoría no encontrada."}, status=404)
            productos = Producto.objects.filter(categoria=categoria)

        elif tipo == "sub":
            subcategoria = Subcategoria.objects.filter(id=grupo_id).first()
            if not subcategoria:
                return Response({"error": "Subcategoría no encontrada."}, status=404)
            productos = Producto.objects.filter(subcategoria=subcategoria)

        elif tipo == "prov":
            proveedor = Proveedor.objects.filter(id=grupo_id).first()
            if not proveedor:
                return Response({"error": "Proveedor no encontrado."}, status=404)
            productos = Producto.objects.filter(proveedor=proveedor)

        # Armar respuesta
        data = []
        for p in productos:
            nuevo_precio = round(p.precio * (1 + porcentaje / 100), 2)
            data.append({
                "codigo": p.codigo,
                "nombre": p.nombre,
                "precio_actual": float(p.precio),
                "precio_nuevo": nuevo_precio
            })

        return Response({
            "tipo": tipo,
            "grupo_id": grupo_id,
            "porcentaje_aplicado": porcentaje,
            "total_afectados": len(data),
            "productos": data
        }, status=200)


class AplicarAjustePreciosView(APIView):
    def post(self, request):
        tipo = request.data.get("tipo")
        grupo_id = request.data.get("id")
        porcentaje = request.data.get("porcentaje")

        if tipo not in ["cat", "sub", "prov"]:
            return Response({"error": "Tipo debe ser 'cat', 'sub' o 'prov'."}, status=400)

        try:
            porcentaje = float(porcentaje)
        except (TypeError, ValueError):
            return Response({"error": "El porcentaje debe ser un número válido."}, status=400)

        if tipo == "cat":
            grupo = Categoria.objects.filter(id=grupo_id).first()
            if not grupo:
                return Response({"error": "Categoría no encontrada."}, status=404)
            productos = Producto.objects.filter(categoria=grupo)

        elif tipo == "sub":
            grupo = Subcategoria.objects.filter(id=grupo_id).first()
            if not grupo:
                return Response({"error": "Subcategoría no encontrada."}, status=404)
            productos = Producto.objects.filter(subcategoria=grupo)

        elif tipo == "prov":
            grupo = Proveedor.objects.filter(id=grupo_id).first()
            if not grupo:
                return Response({"error": "Proveedor no encontrado."}, status=404)
            productos = Producto.objects.filter(proveedor=grupo)

        productos_modificados = 0
        for p in productos:
            nuevo_precio = round(p.precio * (1 + porcentaje / 100), 2)
            if nuevo_precio != p.precio:
                p.precio = nuevo_precio
                p.save()
                productos_modificados += 1

        return Response({
            "mensaje": f"Precios actualizados exitosamente.",
            "tipo": tipo,
            "grupo_id": grupo_id,
            "porcentaje_aplicado": porcentaje,
            "total_modificados": productos_modificados
        }, status=200)

